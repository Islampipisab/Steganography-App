"""
Build Eslam & Farida project timetable Excel matching Project timetable_template.xlsx style.
Output: Downloads/Eslam_Farida_Project_timetable.xlsx
"""
from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TEMPLATE = Path(r"c:\Users\pipis\Downloads\Project timetable_template.xlsx")
OUT = Path(r"c:\Users\pipis\Downloads\Eslam_Farida_Project_timetable.xlsx")

# Light fill for Gantt weeks (similar to thesis blue!30 idea)
FILL_WEEK = PatternFill(fill_type="solid", fgColor="B4C6E7")
FONT_TNR = Font(name="Times New Roman", size=11)
CHECK = "ü"  # same symbol as template


def main() -> None:
    if not TEMPLATE.exists():
        raise SystemExit(f"Template not found: {TEMPLATE}")
    shutil.copy2(TEMPLATE, OUT)
    wb = load_workbook(OUT)

    # ---- Sheet1: calendar-style plan (14 milestones M1–M14) ----
    ws = wb["Sheet1"]
    last_col = 17  # D..Q = 14 months (D=4, Q=17)
    last_letter = get_column_letter(last_col)

    ws["C3"] = "Project  Duration: (14 months) "
    ws["C4"] = "          Number of Phases: Two Phases"

    # Retitle phase rows for Grad II steganography / thesis
    ws["C7"] = "Phase 1 — Research, requirements & design"
    ws["B8"] = "Obj #1"
    ws["C8"] = "Literature review & steganography fundamentals"
    ws["C9"] = "Activity 1: Project kickoff & scope"
    ws["C10"] = "Activity 2: Basics of steganography & requirements"
    ws["C11"] = "Activity 3: Research steganography types (LSB, DCT, hybrid)"
    ws["C12"] = "Chapters 1–2 (Introduction & literature review)"
    ws["C13"] = "Activity 1: First presentation preparation"
    ws["B14"] = "Obj #2"
    ws["C14"] = "System modelling & analysis (Chapter 3)"
    ws["C15"] = "Activity 2: Architecture, workflows & UML diagrams"
    ws["C16"] = "Phase 2 — Implementation, evaluation & thesis completion"
    ws["B17"] = "Obj #3"
    ws["C17"] = "Software implementation (LSB, DCT, hybrid, web UI)"
    ws["C18"] = "Activity 1: Core modules & barcode pipeline"
    ws["C19"] = "Activity 2: Implementation improvements & deployment (e.g. cloud)"
    ws["C20"] = "Activity 3: Testing & quality metrics (PSNR, SSIM)"
    ws["B21"] = "Obj #4"
    ws["C21"] = "Thesis chapters & final deliverables"
    ws["C22"] = "Activity 1: Requirement / cost / time-plan chapters"
    ws["C23"] = "Activity 2: Conclusion & final report assembly"
    ws["C24"] = "Activity 3: Presentation & demo preparation"
    ws["B25"] = "Obj #5"
    ws["C25"] = "Activity 4: Proofreading & submission readiness"
    ws["C26"] = "Activity 5: (Reserve / buffer for revisions)"
    ws["C27"] = "Activity 6: (Reserve / buffer for supervisor feedback)"

    # Clear old milestone markers in calendar area (D8:last_letter27)
    for r in range(8, 28):
        for c in range(4, last_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()

    def mark(row: int, months: list[int]) -> None:
        for m in months:
            col = 3 + m  # M1 -> column 4
            ws.cell(row, col).value = CHECK
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

    # Month indices 1..14 aligned with user's LaTeX Gantt
    mark(9, [1])
    mark(10, [2])
    mark(11, [3])
    mark(12, [2, 3, 4, 5, 6])
    mark(13, [6])
    mark(14, [4, 5, 6, 7, 8, 9, 10, 11])
    mark(15, [9, 10, 11])
    mark(18, [7])
    mark(19, [7, 12])
    mark(20, [13, 14])
    mark(22, [10, 11])
    mark(23, [12, 13, 14])
    mark(24, [14])
    mark(25, [13, 14])
    mark(26, [13])
    mark(27, [14])

    # Optional: trim header row 6 to M1..M14 only (hide M15)
    ws.cell(6, 18).value = None  # R6 was M15 in full template; column 18 = R

    # ---- Sheet2: task list + week columns with fills ----
    w2 = wb["Sheet2"]
    # Clear data rows 3 onward (keep header style in rows 1-2)
    for r in range(3, w2.max_row + 1):
        for c in range(1, 17):
            cell = w2.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()

    tasks: list[tuple[str, str | None, str | None, list[int]]] = [
        ("Project kickoff meeting", "Eslam", "Farida", [1]),
        ("Basics of steganography & requirements", "Eslam", "Farida", [2]),
        ("Start Chapter 1 (Introduction)", "Eslam", "Farida", [2, 3]),
        ("Research steganography types", "Eslam", "Farida", [3]),
        ("Start Chapter 2 (Literature review)", "Eslam", "Farida", [4, 5]),
        ("Finalize literature review", "Eslam", None, [6]),
        ("First presentation preparation", "Eslam", "Farida", [6]),
        ("Design first program module (LSB)", "Eslam", None, [7]),
        ("System architecture & workflow draft", "Eslam", "Farida", [7]),
        ("Start Chapter 3 (diagrams)", "Eslam", "Farida", [8, 9, 10]),
        ("Block / flow / activity / class / sequence diagrams", "Eslam", "Farida", [9, 10, 11]),
        ("Requirement analysis (Chapter 3)", "Farida", None, [10, 11]),
        ("Cost analysis (Chapter 4)", "Farida", None, [11]),
        ("Time plan (Chapter 5)", "Eslam", "Farida", [11]),
        ("Implementation improvements", "Eslam", "Farida", [12]),
        ("Start final chapter (Conclusion)", "Farida", None, [12, 13, 14]),
        ("Testing & quality evaluation", "Eslam", "Farida", [13, 14]),
        ("Final report assembly & proofreading", "Eslam", "Farida", [13, 14]),
        ("Presentation & demo preparation", "Eslam", "Farida", [14]),
    ]

    w2["D1"] = "Weeks (M1–M14)"
    for i in range(14):
        w2.cell(2, 4 + i).value = i + 1

    for idx, (name, h1, h2, weeks) in enumerate(tasks, start=3):
        w2.cell(idx, 1).value = name
        w2.cell(idx, 2).value = h1
        w2.cell(idx, 3).value = h2
        for w in weeks:
            c = 3 + w
            cell = w2.cell(idx, c)
            cell.value = CHECK
            cell.fill = FILL_WEEK
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
